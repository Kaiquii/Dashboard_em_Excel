from pathlib import Path
import re
import unicodedata

import openpyxl
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


INPUT = Path(r"C:\Users\kaiqu\Downloads\CQ Produto Acabado - Procytrat 2026.xlsm")
OUTPUT = Path("CQ Produto Acabado - Procytrat 2026 - Dashboard2-profissional-sem-simulados.xlsm")

ROWS_PER_PRODUCT = 500
HELPER_FIRST_ROW = 2
EXCLUDED = {"dashboard", "dashboard2", "menu", "padrao", "planilha1", "planilha3"}

COLORS = {
    "bg": "F3F6FA",
    "nav": "17212B",
    "nav2": "213043",
    "surface": "FFFFFF",
    "surface2": "F8FAFC",
    "line": "D7DEE8",
    "line2": "E7EDF4",
    "ink": "111827",
    "muted": "667085",
    "blue": "2563EB",
    "blue2": "EAF1FF",
    "teal": "0F766E",
    "teal2": "E7F7F5",
    "green": "16A34A",
    "green2": "E8F8EF",
    "amber": "D97706",
    "amber2": "FFF4D6",
    "rose": "BE123C",
    "rose2": "FFE4E8",
    "purple": "7C3AED",
    "purple2": "F1E9FF",
}


def normalize(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def excel_text(value):
    return str(value).replace('"', '""')


def quote_sheet(name):
    return "'" + name.replace("'", "''") + "'"


def product_name(sheet_name):
    return re.sub(r"^procytrat\s+", "", sheet_name, flags=re.I).strip() or sheet_name


def find_header(ws):
    best_row = None
    best_score = -1
    best_values = []
    max_scan_row = min(ws.max_row, 10)
    max_scan_col = min(ws.max_column, 35)

    for row in range(1, max_scan_row + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, max_scan_col + 1)]
        text = normalize(" | ".join(str(v) for v in values if v is not None))
        score = sum(token in text for token in ("data", "recebimento", "lote", "quantidade", "status", "aprovacao", "validade"))
        if score > best_score:
            best_row = row
            best_score = score
            best_values = values

    if best_score < 2:
        return None

    header = {"row": best_row}
    for idx, raw in enumerate(best_values, start=1):
        text = normalize(raw)
        if not text:
            continue
        if ("data" in text or "recebimento" in text) and "validade" not in text:
            header.setdefault("date", idx)
        if "lote" in text:
            header.setdefault("lot", idx)
        if "quantidade" in text:
            header.setdefault("qty", idx)
        if "status" in text or "aprovacao" in text:
            header.setdefault("status", idx)
        if "validade" in text:
            header.setdefault("validity", idx)

    return header if {"date", "lot", "qty"}.issubset(header) else None


def fill(color):
    return PatternFill("solid", fgColor=color)


def border(color=COLORS["line"], style="thin"):
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def merge(ws, ref, value=None, fill_color=None, font=None, alignment=None, border_color=None):
    ws.merge_cells(ref)
    top_left = ref.split(":")[0]
    cell = ws[top_left]
    if value is not None:
        cell.value = value
    if fill_color:
        for row in ws[ref]:
            for item in row:
                item.fill = fill(fill_color)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border_color:
        for row in ws[ref]:
            for item in row:
                item.border = border(border_color)


def style_area(ws, ref, fill_color, border_color=None):
    for row in ws[ref]:
        for cell in row:
            cell.fill = fill(fill_color)
            cell.alignment = Alignment(vertical="center")
            if border_color:
                cell.border = border(border_color)
            else:
                cell.border = Border()


def outline_range(ws, ref, color=COLORS["line"]):
    start, end = ref.split(":")
    min_col, min_row = ws[start].column, ws[start].row
    max_col, max_row = ws[end].column, ws[end].row
    side = Side(style="thin", color=color)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            left = side if col == min_col else Side(style=None)
            right = side if col == max_col else Side(style=None)
            top = side if row == min_row else Side(style=None)
            bottom = side if row == max_row else Side(style=None)
            ws.cell(row, col).border = Border(left=left, right=right, top=top, bottom=bottom)


def panel(ws, ref, title, subtitle=None):
    start, end = ref.split(":")
    min_col, min_row = ws[start].column, ws[start].row
    max_col, max_row = ws[end].column, ws[end].row
    style_area(ws, ref, COLORS["surface"])
    outline_range(ws, ref, COLORS["line"])
    merge(
        ws,
        f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{min_row + 1}",
        title,
        COLORS["nav2"],
        Font(name="Segoe UI", color="FFFFFF", bold=True, size=11),
        Alignment(horizontal="left", vertical="center"),
        COLORS["nav2"],
    )
    ws.cell(min_row, min_col).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if subtitle:
        merge(
            ws,
            f"{get_column_letter(min_col)}{max_row}:{get_column_letter(max_col)}{max_row}",
            subtitle,
            COLORS["surface2"],
            Font(name="Segoe UI", color=COLORS["muted"], size=9),
            Alignment(horizontal="left", vertical="center"),
            COLORS["line"],
        )


def card(ws, ref, title, formula, accent, number_format):
    start, end = ref.split(":")
    min_col, min_row = ws[start].column, ws[start].row
    max_col, max_row = ws[end].column, ws[end].row
    style_area(ws, ref, COLORS["surface"])
    outline_range(ws, ref, COLORS["line"])
    for col in range(min_col, max_col + 1):
        ws.cell(min_row, col).fill = fill(accent)
        ws.cell(min_row, col).border = border(accent)
    merge(
        ws,
        f"{get_column_letter(min_col)}{min_row + 1}:{get_column_letter(max_col)}{min_row + 1}",
        title,
        COLORS["surface"],
        Font(name="Segoe UI", color=COLORS["muted"], bold=True, size=9),
        Alignment(horizontal="left", vertical="center", indent=1),
        COLORS["line"],
    )
    merge(
        ws,
        f"{get_column_letter(min_col)}{min_row + 2}:{get_column_letter(max_col)}{max_row}",
        formula,
        COLORS["surface"],
        Font(name="Segoe UI", color=accent, bold=True, size=19),
        Alignment(horizontal="left", vertical="center", indent=1),
        COLORS["line"],
    )
    value_cell = ws.cell(min_row + 2, min_col)
    value_cell.number_format = number_format


def table_style(ws, header_row, first_col, last_row, last_col, accent):
    for col in range(first_col, last_col + 1):
        c = ws.cell(header_row, col)
        c.fill = fill(accent)
        c.font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border(accent)

    for row in range(header_row + 1, last_row + 1):
        row_fill = COLORS["surface"] if row % 2 else COLORS["surface2"]
        for col in range(first_col, last_col + 1):
            c = ws.cell(row, col)
            c.fill = fill(row_fill)
            c.font = Font(name="Segoe UI", color=COLORS["ink"], size=9)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = border(COLORS["line2"])


def build():
    wb = openpyxl.load_workbook(INPUT, keep_vba=True)

    source_sheets = []
    for item in wb.worksheets:
        if normalize(item.title) in EXCLUDED:
            continue
        header = find_header(item)
        if header:
            source_sheets.append({"sheet": item.title, "label": product_name(item.title), "header": header})

    for name in ("Dashboard2", "dashboard2"):
        if name in wb.sheetnames:
            wb.remove(wb[name])

    ws = wb.create_sheet("Dashboard2", 0)
    wb.active = 0
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "E11"

    widths = {
        "A": 2.5,
        "B": 13,
        "C": 13,
        "D": 2.5,
        "E": 11,
        "F": 11,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 9,
        "P": 7,
        "Q": 17,
        "R": 15,
        "S": 9,
        "T": 10,
        "U": 11,
        "V": 10,
        "W": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(1, 101):
        ws.row_dimensions[row].height = 22
    for row in (1, 2, 3, 5, 6, 7, 8, 9):
        ws.row_dimensions[row].height = 24
    ws.row_dimensions[2].height = 30

    style_area(ws, "A1:W100", COLORS["bg"])
    style_area(ws, "A1:C100", COLORS["nav"])
    style_area(ws, "D1:D100", COLORS["bg"])

    merge(
        ws,
        "B2:C2",
        "PROCYTRAT",
        COLORS["nav"],
        Font(name="Segoe UI", color="FFFFFF", bold=True, size=16),
        Alignment(horizontal="left", vertical="center"),
    )
    merge(
        ws,
        "B3:C3",
        "Controle de qualidade",
        COLORS["nav"],
        Font(name="Segoe UI", color="B8C4D6", size=9),
        Alignment(horizontal="left", vertical="center"),
    )
    merge(
        ws,
        "B6:C7",
        '=COUNTA($E$67:$E$86)',
        COLORS["nav2"],
        Font(name="Segoe UI", color="FFFFFF", bold=True, size=22),
        Alignment(horizontal="center", vertical="center"),
        COLORS["nav2"],
    )
    ws["B6"].number_format = "0"
    merge(
        ws,
        "B8:C8",
        "produtos ativos",
        COLORS["nav2"],
        Font(name="Segoe UI", color="B8C4D6", bold=True, size=9),
        Alignment(horizontal="center", vertical="center"),
        COLORS["nav2"],
    )
    for row, label in [(12, "Vis\u00e3o geral"), (14, "Ranking"), (16, "Mensal"), (18, "Status"), (20, "Produtos")]:
        merge(
            ws,
            f"B{row}:C{row}",
            label,
            COLORS["nav"],
            Font(name="Segoe UI", color="DDE6F3", bold=True, size=10),
            Alignment(horizontal="left", vertical="center", indent=1),
        )

    merge(
        ws,
        "E2:S2",
        "Dashboard de Produto Acabado",
        COLORS["bg"],
        Font(name="Segoe UI", color=COLORS["ink"], bold=True, size=22),
        Alignment(horizontal="left", vertical="center"),
    )
    merge(
        ws,
        "E3:S3",
        '="Atualizado: "&TEXT(NOW(),"dd/mm/aaaa hh:mm")&"  |  Ano selecionado: "&$W$3',
        COLORS["bg"],
        Font(name="Segoe UI", color=COLORS["muted"], size=10),
        Alignment(horizontal="left", vertical="center"),
    )
    merge(
        ws,
        "U2:V2",
        "ANO",
        COLORS["surface"],
        Font(name="Segoe UI", color=COLORS["muted"], bold=True, size=9),
        Alignment(horizontal="center", vertical="center"),
        COLORS["line"],
    )
    merge(
        ws,
        "U3:V3",
        "sele\u00e7\u00e3o",
        COLORS["surface"],
        Font(name="Segoe UI", color=COLORS["muted"], size=9),
        Alignment(horizontal="center", vertical="center"),
        COLORS["line"],
    )
    ws["W2"] = "Selecionado"
    ws["W2"].fill = fill(COLORS["surface"])
    ws["W2"].font = Font(name="Segoe UI", color=COLORS["muted"], bold=True, size=9)
    ws["W2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["W2"].border = border(COLORS["line"])
    ws["W3"] = 2026
    ws["W3"].fill = fill(COLORS["amber2"])
    ws["W3"].font = Font(name="Segoe UI", color=COLORS["ink"], bold=True, size=11)
    ws["W3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["W3"].border = border(COLORS["line"])

    card(ws, "E5:H9", "TOTAL DE LOTES", '=SUMIFS($AJ$2:$AJ$12001,$AG$2:$AG$12001,$W$3)', COLORS["blue"], "0")
    card(ws, "I5:L9", "QUANTIDADE TOTAL", '=SUMIFS($AD$2:$AD$12001,$AJ$2:$AJ$12001,1,$AG$2:$AG$12001,$W$3)', COLORS["teal"], '#,##0 "kg"')
    card(ws, "M5:P9", "LOTES APROVADOS", '=SUMIFS($AF$2:$AF$12001,$AJ$2:$AJ$12001,1,$AG$2:$AG$12001,$W$3)', COLORS["green"], "0")
    card(
        ws,
        "Q5:T9",
        "TAXA DE APROVA\u00c7\u00c3O",
        '=IFERROR(SUMIFS($AF$2:$AF$12001,$AJ$2:$AJ$12001,1,$AG$2:$AG$12001,$W$3)/SUMIFS($AJ$2:$AJ$12001,$AG$2:$AG$12001,$W$3),0)',
        COLORS["amber"],
        "0.0%",
    )
    card(ws, "U5:W9", "PRODUTOS ATIVOS", '=COUNTA($E$67:$E$86)', COLORS["purple"], "0")

    panel(ws, "E12:N27", "Top 10 por quantidade")
    panel(ws, "P12:W27", "Ranking detalhado")
    panel(ws, "E29:N45", "Produ\u00e7\u00e3o mensal")
    panel(ws, "P29:W45", "Detalhe mensal")
    panel(ws, "E47:K61", "Distribui\u00e7\u00e3o de status")
    panel(ws, "M47:Q61", "Status dos lotes")
    panel(ws, "S47:W61", "Pr\u00f3ximos vencimentos")
    panel(ws, "E64:W88", "Resumo por produto")

    top_headers = ["#", "Produto", "Quantidade (kg)", "Lotes", "% Aprov.", "Situa\u00e7\u00e3o"]
    for idx, text in enumerate(top_headers, start=16):
        ws.cell(15, idx).value = text
    table_style(ws, 15, 16, 25, 21, COLORS["blue"])
    for row in range(16, 26):
        rank = row - 15
        ws.cell(row, 16).value = rank
        ws.cell(row, 17).value = f'=IFERROR(INDEX($E$67:$E$86,MATCH(LARGE($G$67:$G$86,P{row}),$G$67:$G$86,0)),"")'
        ws.cell(row, 18).value = f'=IFERROR(LARGE($G$67:$G$86,P{row}),"")'
        ws.cell(row, 19).value = f'=IFERROR(INDEX($F$67:$F$86,MATCH(Q{row},$E$67:$E$86,0)),"")'
        ws.cell(row, 20).value = f'=IFERROR(INDEX($I$67:$I$86,MATCH(Q{row},$E$67:$E$86,0)),"")'
        ws.cell(row, 21).value = f'=IF(Q{row}="","",IF(T{row}>=0.95,"Excelente",IF(T{row}>=0.8,"Aten\u00e7\u00e3o","Cr\u00edtico")))'
        ws.cell(row, 18).number_format = '#,##0 "kg"'
        ws.cell(row, 20).number_format = "0.0%"

    month_headers = ["M\u00eas", "Lotes", "Quantidade (kg)", "Aprovados", "% Aprov."]
    for idx, text in enumerate(month_headers, start=16):
        ws.cell(32, idx).value = text
    table_style(ws, 32, 16, 44, 20, COLORS["teal"])
    months = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    for month_no, month in enumerate(months, start=1):
        row = 32 + month_no
        ws.cell(row, 16).value = month
        ws.cell(row, 17).value = f'=SUMIFS($AJ$2:$AJ$12001,$AG$2:$AG$12001,$W$3,$AH$2:$AH$12001,{month_no})'
        ws.cell(row, 18).value = f'=SUMIFS($AD$2:$AD$12001,$AJ$2:$AJ$12001,1,$AG$2:$AG$12001,$W$3,$AH$2:$AH$12001,{month_no})'
        ws.cell(row, 19).value = f'=SUMIFS($AF$2:$AF$12001,$AJ$2:$AJ$12001,1,$AG$2:$AG$12001,$W$3,$AH$2:$AH$12001,{month_no})'
        ws.cell(row, 20).value = f'=IFERROR(S{row}/Q{row},0)'
        ws.cell(row, 18).number_format = '#,##0 "kg"'
        ws.cell(row, 20).number_format = "0.0%"

    status_headers = ["Status", "Lotes", "%"]
    for idx, text in enumerate(status_headers, start=13):
        ws.cell(50, idx).value = text
    table_style(ws, 50, 13, 53, 15, COLORS["green"])
    status_items = [
        ("Aprovado", '=SUMIFS($AF$2:$AF$12001,$AJ$2:$AJ$12001,1,$AG$2:$AG$12001,$W$3)'),
        ("Em andamento", '=SUMIFS($AJ$2:$AJ$12001,$AE$2:$AE$12001,"*andamento*",$AG$2:$AG$12001,$W$3)'),
        ("Reprovado", '=SUMIFS($AJ$2:$AJ$12001,$AE$2:$AE$12001,"*reprovado*",$AG$2:$AG$12001,$W$3)'),
    ]
    for row, (label, formula) in enumerate(status_items, start=51):
        ws.cell(row, 13).value = label
        ws.cell(row, 14).value = formula
        ws.cell(row, 15).value = f'=IFERROR(N{row}/$E$7,0)'
        ws.cell(row, 15).number_format = "0.0%"
    for row, color in zip(range(51, 54), [COLORS["green"], COLORS["amber"], COLORS["rose"]]):
        ws.cell(row, 13).fill = fill(color)
        ws.cell(row, 13).font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=9)
        ws.cell(row, 13).alignment = Alignment(horizontal="left", vertical="center", indent=1)

    due_headers = ["#", "Produto", "Lote", "Validade", "Dias"]
    for idx, text in enumerate(due_headers, start=19):
        ws.cell(50, idx).value = text
    table_style(ws, 50, 19, 60, 23, COLORS["rose"])
    for row in range(51, 61):
        rank = row - 50
        ws.cell(row, 19).value = rank
        ws.cell(row, 20).value = f'=IFERROR(INDEX($AA$2:$AA$12001,MATCH(SMALL($AM$2:$AM$12001,S{row}),$AM$2:$AM$12001,0)),"")'
        ws.cell(row, 21).value = f'=IFERROR(INDEX($AC$2:$AC$12001,MATCH(SMALL($AM$2:$AM$12001,S{row}),$AM$2:$AM$12001,0)),"")'
        ws.cell(row, 22).value = f'=IFERROR(INDEX($AL$2:$AL$12001,MATCH(SMALL($AM$2:$AM$12001,S{row}),$AM$2:$AM$12001,0)),"")'
        ws.cell(row, 23).value = f'=IF(V{row}="","",V{row}-TODAY())'
        ws.cell(row, 22).number_format = "dd/mm/yyyy"
        ws.cell(row, 23).number_format = "0"

    product_headers = ["Produto", "Lotes", "Quantidade (kg)", "Aprovados", "% Aprov.", "Rank", "Situa\u00e7\u00e3o"]
    for idx, text in enumerate(product_headers, start=5):
        ws.cell(66, idx).value = text
    table_style(ws, 66, 5, 86, 11, COLORS["purple"])
    for idx, source in enumerate(source_sheets, start=67):
        label = excel_text(source["label"])
        ws.cell(idx, 5).value = source["label"]
        ws.cell(idx, 6).value = f'=SUMIFS($AJ$2:$AJ$12001,$AA$2:$AA$12001,E{idx},$AG$2:$AG$12001,$W$3)'
        ws.cell(idx, 7).value = f'=SUMIFS($AD$2:$AD$12001,$AJ$2:$AJ$12001,1,$AA$2:$AA$12001,E{idx},$AG$2:$AG$12001,$W$3)'
        ws.cell(idx, 8).value = f'=SUMIFS($AF$2:$AF$12001,$AJ$2:$AJ$12001,1,$AA$2:$AA$12001,E{idx},$AG$2:$AG$12001,$W$3)'
        ws.cell(idx, 9).value = f'=IFERROR(H{idx}/F{idx},0)'
        ws.cell(idx, 10).value = f'=IF(G{idx}>0,RANK(G{idx},$G$67:$G$86,0),"")'
        ws.cell(idx, 11).value = f'=IF(F{idx}=0,"",IF(I{idx}>=0.95,"Excelente",IF(I{idx}>=0.8,"Aten\u00e7\u00e3o","Cr\u00edtico")))'
        ws.cell(idx, 7).number_format = '#,##0 "kg"'
        ws.cell(idx, 9).number_format = "0.0%"

    helper_headers = [
        "Produto",
        "Data",
        "Lote",
        "Quantidade",
        "Status",
        "Aprovado",
        "Ano",
        "M\u00eas",
        "Ordena\u00e7\u00e3o",
        "Linha v\u00e1lida",
        "Flag",
        "Validade calc.",
        "Vencimento ordenado",
    ]
    for idx, text in enumerate(helper_headers, start=27):
        ws.cell(1, idx).value = text
        ws.cell(1, idx).font = Font(name="Segoe UI", bold=True)

    current_row = HELPER_FIRST_ROW
    for source in source_sheets:
        sheet_ref = quote_sheet(source["sheet"])
        h = source["header"]
        date_col = get_column_letter(h["date"])
        lot_col = get_column_letter(h["lot"])
        qty_col = get_column_letter(h["qty"])
        status_col = get_column_letter(h["status"]) if "status" in h else None
        validity_col = get_column_letter(h["validity"]) if "validity" in h else None
        first_source_row = h["row"] + 1
        label = excel_text(source["label"])

        for offset in range(ROWS_PER_PRODUCT):
            src_row = first_source_row + offset
            date_ref = f"{sheet_ref}!{date_col}{src_row}"
            lot_ref = f"{sheet_ref}!{lot_col}{src_row}"
            qty_ref = f"{sheet_ref}!{qty_col}{src_row}"
            status_ref = f"{sheet_ref}!{status_col}{src_row}" if status_col else '""'
            validity_ref = f"{sheet_ref}!{validity_col}{src_row}" if validity_col else '""'
            product_cell = f"AA{current_row}"
            date_cell = f"AB{current_row}"
            status_cell = f"AE{current_row}"
            valid_cell = f"AJ{current_row}"
            validity_cell = f"AL{current_row}"

            ws.cell(current_row, 27).value = f'=IF(COUNTA({date_ref},{lot_ref},{qty_ref})>0,"{label}","")'
            ws.cell(current_row, 28).value = f'=IFERROR(IF(ISNUMBER({date_ref}),{date_ref},DATEVALUE({date_ref})),"")'
            ws.cell(current_row, 29).value = f'=IF({product_cell}="","",{lot_ref})'
            ws.cell(current_row, 30).value = (
                f'=IF({product_cell}="","",IFERROR(1*{qty_ref},IFERROR(NUMBERVALUE({qty_ref},".",","),'
                f'IFERROR(NUMBERVALUE({qty_ref},",","."),0))))'
            )
            ws.cell(current_row, 31).value = f'=IF({product_cell}="","",{status_ref})'
            ws.cell(current_row, 32).value = f'=IF({product_cell}="",0,--ISNUMBER(SEARCH("aprovado",{status_cell})))'
            ws.cell(current_row, 33).value = f'=IF({date_cell}="","",YEAR({date_cell}))'
            ws.cell(current_row, 34).value = f'=IF({date_cell}="","",MONTH({date_cell}))'
            ws.cell(current_row, 35).value = f'=IF({date_cell}="",0,{date_cell}+ROW()/1000000)'
            ws.cell(current_row, 36).value = f'=--({product_cell}<>"")'
            ws.cell(current_row, 37).value = 0
            ws.cell(current_row, 38).value = (
                f'=IF({product_cell}="","",IFERROR(IF(ISNUMBER({validity_ref}),{validity_ref},'
                f'IF(ISNUMBER(SEARCH("12",{validity_ref})),EDATE({date_cell},12),'
                f'IF(ISNUMBER(SEARCH("6",{validity_ref})),EDATE({date_cell},6),'
                f'IF(ISNUMBER(SEARCH("3",{validity_ref})),EDATE({date_cell},3),DATEVALUE({validity_ref}))))),""))'
            )
            ws.cell(current_row, 39).value = f'=IF(AND({valid_cell}=1,{validity_cell}<>"",{validity_cell}>=TODAY(),AG{current_row}=$W$3),{validity_cell}+ROW()/1000000,"")'
            current_row += 1

    for idx, year in enumerate(range(2022, 2031), start=2):
        ws.cell(idx, 40).value = year
    validation = DataValidation(type="list", formula1="$AN$2:$AN$10", allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(ws["W3"])

    ws.conditional_formatting.add("G67:G86", DataBarRule(start_type="min", end_type="max", color=COLORS["blue"], showValue=True))
    ws.conditional_formatting.add("R16:R25", DataBarRule(start_type="min", end_type="max", color=COLORS["green"], showValue=True))
    ws.conditional_formatting.add("R33:R44", DataBarRule(start_type="min", end_type="max", color=COLORS["teal"], showValue=True))
    ws.conditional_formatting.add("I67:I86", CellIsRule(operator="lessThan", formula=["0.8"], fill=fill(COLORS["rose2"])))
    ws.conditional_formatting.add("I67:I86", CellIsRule(operator="greaterThanOrEqual", formula=["0.95"], fill=fill(COLORS["green2"])))
    ws.conditional_formatting.add("U16:U25", FormulaRule(formula=['U16="Excelente"'], fill=fill(COLORS["green2"])))
    ws.conditional_formatting.add("U16:U25", FormulaRule(formula=['U16="Aten\u00e7\u00e3o"'], fill=fill(COLORS["amber2"])))
    ws.conditional_formatting.add("U16:U25", FormulaRule(formula=['U16="Cr\u00edtico"'], fill=fill(COLORS["rose2"])))

    top_chart = BarChart()
    top_chart.type = "bar"
    top_chart.title = None
    top_chart.add_data(Reference(ws, min_col=18, min_row=15, max_row=25), titles_from_data=True)
    top_chart.set_categories(Reference(ws, min_col=17, min_row=16, max_row=25))
    top_chart.x_axis.title = "kg"
    top_chart.y_axis.title = "Produto"
    top_chart.x_axis.delete = False
    top_chart.y_axis.delete = False
    top_chart.height = 7.4
    top_chart.width = 18.4
    top_chart.style = 10
    top_chart.legend = None
    if top_chart.series:
        top_chart.series[0].graphicalProperties.solidFill = COLORS["blue"]
        top_chart.series[0].graphicalProperties.line.solidFill = COLORS["blue"]
    ws.add_chart(top_chart, "E14")

    month_chart = BarChart()
    month_chart.type = "col"
    month_chart.title = None
    month_chart.add_data(Reference(ws, min_col=18, min_row=32, max_row=44), titles_from_data=True)
    month_chart.set_categories(Reference(ws, min_col=16, min_row=33, max_row=44))
    month_chart.y_axis.title = "kg"
    month_chart.x_axis.title = "M\u00eas"
    month_chart.x_axis.delete = False
    month_chart.y_axis.delete = False
    month_chart.height = 7.4
    month_chart.width = 18.4
    month_chart.style = 10
    month_chart.legend = None
    if month_chart.series:
        month_chart.series[0].graphicalProperties.solidFill = COLORS["teal"]
        month_chart.series[0].graphicalProperties.line.solidFill = COLORS["teal"]
    ws.add_chart(month_chart, "E31")

    status_chart = DoughnutChart()
    status_chart.title = None
    status_chart.add_data(Reference(ws, min_col=14, min_row=50, max_row=53), titles_from_data=True)
    status_chart.set_categories(Reference(ws, min_col=13, min_row=51, max_row=53))
    status_chart.holeSize = 58
    status_chart.height = 5.8
    status_chart.width = 7.8
    status_chart.style = 10
    status_chart.legend = None
    ws.add_chart(status_chart, "F50")

    # Keep helper data available for formulas but out of the dashboard canvas.
    for col in range(24, 41):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=23):
        for cell in row:
            if cell.font == Font():
                cell.font = Font(name="Segoe UI", color=COLORS["ink"], size=9)
            if cell.alignment is None:
                cell.alignment = Alignment(vertical="center")

    for row in range(16, 26):
        ws.cell(row, 16).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 18).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, 20).alignment = Alignment(horizontal="right", vertical="center")
    for row in range(33, 45):
        ws.cell(row, 18).alignment = Alignment(horizontal="right", vertical="center")
    for row in range(51, 55):
        ws.cell(row, 14).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, 15).alignment = Alignment(horizontal="right", vertical="center")
    for row in range(51, 61):
        ws.cell(row, 19).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 22).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 23).alignment = Alignment(horizontal="right", vertical="center")

    ws.protection.sheet = False
    ws.protection.objects = False
    ws.protection.scenarios = False
    ws.print_area = "A1:W99"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(OUTPUT)
    return OUTPUT, len(source_sheets)


if __name__ == "__main__":
    output, count = build()
    print(f"Arquivo gerado: {output.resolve()}")
    print(f"Abas de produto usadas: {count}")
