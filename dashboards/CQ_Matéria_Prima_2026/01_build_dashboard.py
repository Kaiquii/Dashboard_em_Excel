from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
import json
import re
import unicodedata

import openpyxl
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]
DASHBOARD_ID = SCRIPT_DIR.name
CONFIG = json.loads((SCRIPT_DIR / "config.json").read_text(encoding="utf-8"))
INPUT = PROJECT_ROOT / CONFIG["entrada"]
WORK_DIR = PROJECT_ROOT / ".tmp" / DASHBOARD_ID
OUTPUT = WORK_DIR / "CQ Materia Prima - 2026 - Dashboard-openpyxl.xlsm"

DEFAULT_YEAR = 2026
ROWS_PER_MATERIAL = 500
HELPER_FIRST_ROW = 2
EXCLUDED = {"dashboard", "dashboard2", "menu", "modelo", "padrao", "planilha1"}

COLORS = {
    "bg": "F3F6FA",
    "nav": "17212B",
    "nav2": "213043",
    "surface": "FFFFFF",
    "surface2": "F8FAFC",
    "line": "D7DEE8",
    "line2": "E7EDF4",
    "ink": "111827",
    "muted": "667085",
    "blue": "2563EB",
    "blue2": "EAF1FF",
    "teal": "0F766E",
    "green": "16A34A",
    "green2": "E8F8EF",
    "amber": "D97706",
    "amber2": "FFF4D6",
    "rose": "BE123C",
    "rose2": "FFE4E8",
    "purple": "7C3AED",
}


def normalize(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def excel_text(value):
    return str(value).replace('"', '""')


def quote_sheet(name):
    return "'" + name.replace("'", "''") + "'"


def classify_header(value):
    text = normalize(value)
    if "aprov" in text or text == "status":
        return "approval"
    if "fornecedor" in text or "forenecedor" in text:
        return "supplier"
    if "receb" in text or "rceb" in text or text == "data":
        return "received_date"
    if "fabric" in text:
        return "manufacturing_date"
    if "lote" in text:
        return "lot"
    if "quantidade" in text:
        return "qty"
    if "validade" in text:
        return "validity"
    if "observ" in text:
        return "notes"
    if "responsavel" in text:
        return "owner"
    return None


def parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_number(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    for decimal, thousands in ((",", "."), (".", ",")):
        cleaned = text.replace(thousands, "").replace(decimal, ".")
        try:
            return float(cleaned)
        except ValueError:
            pass
    return 0.0


def is_approved(value):
    return "aprovado" in normalize(value)


def status_label(value):
    text = normalize(value)
    if "aprovado" in text:
        return "Aprovado"
    if "reprov" in text:
        return "Reprovado"
    if text:
        return "Em andamento"
    return "Sem status"


def fill(color):
    return PatternFill("solid", fgColor=color)


def border(color=COLORS["line"], style="thin"):
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def merge(ws, ref, value=None, fill_color=None, font=None, alignment=None, border_color=None):
    ws.merge_cells(ref)
    cell = ws[ref.split(":")[0]]
    if value is not None:
        cell.value = value
    if fill_color:
        for row in ws[ref]:
            for item in row:
                item.fill = fill(fill_color)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border_color:
        for row in ws[ref]:
            for item in row:
                item.border = border(border_color)


def style_area(ws, ref, fill_color, border_color=None):
    for row in ws[ref]:
        for cell in row:
            cell.fill = fill(fill_color)
            cell.alignment = Alignment(vertical="center")
            cell.border = border(border_color) if border_color else Border()


def outline_range(ws, ref, color=COLORS["line"]):
    start, end = ref.split(":")
    min_col, min_row = ws[start].column, ws[start].row
    max_col, max_row = ws[end].column, ws[end].row
    side = Side(style="thin", color=color)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).border = Border(
                left=side if col == min_col else Side(style=None),
                right=side if col == max_col else Side(style=None),
                top=side if row == min_row else Side(style=None),
                bottom=side if row == max_row else Side(style=None),
            )


def panel(ws, ref, title):
    start, end = ref.split(":")
    min_col, min_row = ws[start].column, ws[start].row
    max_col = ws[end].column
    style_area(ws, ref, COLORS["surface"])
    outline_range(ws, ref)
    merge(
        ws,
        f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{min_row + 1}",
        title,
        COLORS["nav2"],
        Font(name="Segoe UI", color="FFFFFF", bold=True, size=11),
        Alignment(horizontal="left", vertical="center", indent=1),
        COLORS["nav2"],
    )


def card(ws, ref, title, value, accent, number_format):
    start, end = ref.split(":")
    min_col, min_row = ws[start].column, ws[start].row
    max_col, max_row = ws[end].column, ws[end].row
    style_area(ws, ref, COLORS["surface"])
    outline_range(ws, ref)
    for col in range(min_col, max_col + 1):
        ws.cell(min_row, col).fill = fill(accent)
        ws.cell(min_row, col).border = border(accent)
    merge(
        ws,
        f"{get_column_letter(min_col)}{min_row + 1}:{get_column_letter(max_col)}{min_row + 1}",
        title,
        COLORS["surface"],
        Font(name="Segoe UI", color=COLORS["muted"], bold=True, size=9),
        Alignment(horizontal="left", vertical="center", indent=1),
        COLORS["line"],
    )
    merge(
        ws,
        f"{get_column_letter(min_col)}{min_row + 2}:{get_column_letter(max_col)}{max_row}",
        value,
        COLORS["surface"],
        Font(name="Segoe UI", color=accent, bold=True, size=19),
        Alignment(horizontal="left", vertical="center", indent=1),
        COLORS["line"],
    )
    ws.cell(min_row + 2, min_col).number_format = number_format


def table_style(ws, header_row, first_col, last_row, last_col, accent):
    for col in range(first_col, last_col + 1):
        c = ws.cell(header_row, col)
        c.fill = fill(accent)
        c.font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border(accent)

    for row in range(header_row + 1, last_row + 1):
        row_fill = COLORS["surface"] if row % 2 else COLORS["surface2"]
        for col in range(first_col, last_col + 1):
            c = ws.cell(row, col)
            c.fill = fill(row_fill)
            c.font = Font(name="Segoe UI", color=COLORS["ink"], size=9)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = border(COLORS["line2"])


def find_header(ws):
    best_row = None
    best_score = -1
    for row in range(1, min(ws.max_row, 12) + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        text = normalize(" | ".join(str(v) for v in values if v not in (None, "")))
        score = sum(token in text for token in ("aprov", "status", "fornecedor", "receb", "fabric", "lote", "quantidade", "validade"))
        if score > best_score:
            best_score = score
            best_row = row
    if best_score < 3:
        return None

    mapping = {"row": best_row}
    for col in range(1, ws.max_column + 1):
        key = classify_header(ws.cell(best_row, col).value)
        if key and key not in mapping:
            mapping[key] = col
    return mapping if {"approval", "received_date", "qty", "validity"}.issubset(mapping) else None


def collect_records(wb):
    records = []
    materials = []
    for ws in wb.worksheets:
        if normalize(ws.title) in EXCLUDED:
            continue
        header = find_header(ws)
        if not header:
            continue
        materials.append(ws.title)
        for row in range(header["row"] + 1, ws.max_row + 1):
            common = [ws.cell(row, header[key]).value for key in ("approval", "received_date", "qty", "validity") if key in header]
            if "supplier" in header:
                common.append(ws.cell(row, header["supplier"]).value)
            if "lot" in header:
                common.append(ws.cell(row, header["lot"]).value)
            if not any(value not in (None, "") for value in common):
                continue

            supplier = ws.cell(row, header["supplier"]).value if "supplier" in header else ""
            lot = ws.cell(row, header["lot"]).value if "lot" in header else ""
            received = parse_date(ws.cell(row, header["received_date"]).value)
            validity = parse_date(ws.cell(row, header["validity"]).value)
            approval = ws.cell(row, header["approval"]).value
            qty = parse_number(ws.cell(row, header["qty"]).value)
            records.append(
                {
                    "material": ws.title,
                    "supplier": str(supplier or "").strip(),
                    "lot": str(lot or "").strip(),
                    "received": received,
                    "validity": validity,
                    "approval": str(approval or "").strip(),
                    "approved": is_approved(approval),
                    "qty": qty,
                }
            )
    return records, materials


def find_source_sheets(wb):
    source_sheets = []
    for ws in wb.worksheets:
        if normalize(ws.title) in EXCLUDED:
            continue
        header = find_header(ws)
        if header:
            source_sheets.append({"sheet": ws.title, "label": ws.title, "header": header})
    return source_sheets


def summarize(records, materials, year):
    year_records = [item for item in records if item["received"] and item["received"].year == year]
    approved = sum(1 for item in year_records if item["approved"])
    total_qty = sum(item["qty"] for item in year_records)
    suppliers = {item["supplier"] for item in year_records if item["supplier"]}
    active_materials = {item["material"] for item in year_records}

    by_supplier = defaultdict(lambda: {"qty": 0.0, "lots": 0, "approved": 0, "materials": set()})
    by_material = defaultdict(lambda: {"qty": 0.0, "lots": 0, "approved": 0, "suppliers": set()})
    monthly = {month: {"qty": 0.0, "lots": 0, "approved": 0} for month in range(1, 13)}
    status = Counter()

    for item in year_records:
        material = item["material"]
        if item["supplier"]:
            supplier = item["supplier"]
            by_supplier[supplier]["qty"] += item["qty"]
            by_supplier[supplier]["lots"] += 1
            by_supplier[supplier]["approved"] += 1 if item["approved"] else 0
            by_supplier[supplier]["materials"].add(material)

        by_material[material]["qty"] += item["qty"]
        by_material[material]["lots"] += 1
        by_material[material]["approved"] += 1 if item["approved"] else 0
        if item["supplier"]:
            by_material[material]["suppliers"].add(item["supplier"])

        month = item["received"].month
        monthly[month]["qty"] += item["qty"]
        monthly[month]["lots"] += 1
        monthly[month]["approved"] += 1 if item["approved"] else 0
        status[status_label(item["approval"])] += 1

    today = date.today()
    due = [
        item
        for item in year_records
        if item["validity"] and item["validity"] >= today
    ]
    due.sort(key=lambda item: item["validity"])

    for material in materials:
        by_material.setdefault(material, {"qty": 0.0, "lots": 0, "approved": 0, "suppliers": set()})

    return {
        "records": year_records,
        "total_lots": len(year_records),
        "total_qty": total_qty,
        "approved": approved,
        "approval_rate": approved / len(year_records) if year_records else 0,
        "suppliers": suppliers,
        "active_materials": active_materials,
        "by_supplier": by_supplier,
        "by_material": by_material,
        "monthly": monthly,
        "status": status,
        "due": due,
    }


def write_dashboard(wb, summary, materials):
    for name in ("Dashboard2", "dashboard2"):
        if name in wb.sheetnames:
            wb.remove(wb[name])

    ws = wb.create_sheet("Dashboard2", 0)
    wb.active = 0
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "E11"

    widths = {
        "A": 2.5, "B": 13, "C": 13, "D": 2.5,
        "E": 11, "F": 11, "G": 12, "H": 12, "I": 12, "J": 12, "K": 12,
        "L": 12, "M": 12, "N": 12, "O": 9, "P": 7, "Q": 17, "R": 15,
        "S": 10, "T": 10, "U": 11, "V": 10, "W": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(1, 132):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[2].height = 30

    style_area(ws, "A1:W132", COLORS["bg"])
    style_area(ws, "A1:C132", COLORS["nav"])
    style_area(ws, "D1:D132", COLORS["bg"])

    merge(ws, "B2:C2", "MATÉRIA PRIMA", COLORS["nav"], Font(name="Segoe UI", color="FFFFFF", bold=True, size=14), Alignment(horizontal="left", vertical="center"))
    merge(ws, "B3:C3", "Fornecedores", COLORS["nav"], Font(name="Segoe UI", color="B8C4D6", size=9), Alignment(horizontal="left", vertical="center"))
    merge(ws, "B6:C7", len(summary["active_materials"]), COLORS["nav2"], Font(name="Segoe UI", color="FFFFFF", bold=True, size=22), Alignment(horizontal="center", vertical="center"), COLORS["nav2"])
    merge(ws, "B8:C8", "matérias ativas", COLORS["nav2"], Font(name="Segoe UI", color="B8C4D6", bold=True, size=9), Alignment(horizontal="center", vertical="center"), COLORS["nav2"])
    for row, label in [(12, "Visão geral"), (14, "Fornecedores"), (16, "Mensal"), (18, "Status"), (20, "Vencimentos")]:
        merge(ws, f"B{row}:C{row}", label, COLORS["nav"], Font(name="Segoe UI", color="DDE6F3", bold=True, size=10), Alignment(horizontal="left", vertical="center", indent=1))

    merge(ws, "E2:S2", "Dashboard de Matéria Prima", COLORS["bg"], Font(name="Segoe UI", color=COLORS["ink"], bold=True, size=22), Alignment(horizontal="left", vertical="center"))
    merge(ws, "E3:S3", f"Base: fornecedores | Ano base: {DEFAULT_YEAR}", COLORS["bg"], Font(name="Segoe UI", color=COLORS["muted"], size=10), Alignment(horizontal="left", vertical="center"))
    merge(ws, "U2:V2", "ANO", COLORS["surface"], Font(name="Segoe UI", color=COLORS["muted"], bold=True, size=9), Alignment(horizontal="center", vertical="center"), COLORS["line"])
    merge(ws, "U3:V3", "base", COLORS["surface"], Font(name="Segoe UI", color=COLORS["muted"], size=9), Alignment(horizontal="center", vertical="center"), COLORS["line"])
    ws["W2"] = "Selecionado"
    ws["W3"] = DEFAULT_YEAR
    for cell in ("W2", "W3"):
        ws[cell].fill = fill(COLORS["surface"] if cell == "W2" else COLORS["amber2"])
        ws[cell].font = Font(name="Segoe UI", color=COLORS["ink"] if cell == "W3" else COLORS["muted"], bold=True, size=10)
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].border = border(COLORS["line"])

    card(ws, "E5:H9", "RECEBIMENTOS", summary["total_lots"], COLORS["blue"], "0")
    card(ws, "I5:L9", "QUANTIDADE TOTAL", summary["total_qty"], COLORS["teal"], '#,##0 "kg"')
    card(ws, "M5:P9", "FORNECEDORES", len(summary["suppliers"]), COLORS["green"], "0")
    card(ws, "Q5:T9", "TAXA DE APROVAÇÃO", summary["approval_rate"], COLORS["amber"], "0.0%")
    card(ws, "U5:W9", "MATÉRIAS ATIVAS", len(summary["active_materials"]), COLORS["purple"], "0")

    panel(ws, "E12:N27", "Top 10 fornecedores por quantidade")
    panel(ws, "P12:W27", "Ranking de fornecedores")
    panel(ws, "E29:N45", "Recebimentos mensais")
    panel(ws, "P29:W45", "Detalhe mensal")
    panel(ws, "E47:K61", "Distribuição de aprovação")
    panel(ws, "M47:Q61", "Status de aprovação")
    panel(ws, "S47:W61", "Próximos vencimentos")
    panel(ws, "E64:W128", "Resumo por matéria-prima")

    supplier_headers = ["#", "Fornecedor", "Quantidade (kg)", "Receb.", "% Aprov.", "Matérias"]
    for col, text in enumerate(supplier_headers, start=16):
        ws.cell(15, col).value = text
    table_style(ws, 15, 16, 25, 21, COLORS["blue"])
    top_suppliers = sorted(summary["by_supplier"].items(), key=lambda item: item[1]["qty"], reverse=True)[:10]
    for row in range(16, 26):
        idx = row - 16
        ws.cell(row, 16).value = idx + 1
        if idx < len(top_suppliers):
            supplier, data = top_suppliers[idx]
            lots = data["lots"]
            ws.cell(row, 17).value = supplier
            ws.cell(row, 18).value = data["qty"]
            ws.cell(row, 19).value = lots
            ws.cell(row, 20).value = data["approved"] / lots if lots else 0
            ws.cell(row, 21).value = len(data["materials"])
        ws.cell(row, 18).number_format = '#,##0 "kg"'
        ws.cell(row, 20).number_format = "0.0%"

    months = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    month_headers = ["Mês", "Receb.", "Quantidade (kg)", "Aprovados", "% Aprov."]
    for col, text in enumerate(month_headers, start=16):
        ws.cell(32, col).value = text
    table_style(ws, 32, 16, 44, 20, COLORS["teal"])
    for month_no, month in enumerate(months, start=1):
        row = 32 + month_no
        data = summary["monthly"][month_no]
        ws.cell(row, 16).value = month
        ws.cell(row, 17).value = data["lots"]
        ws.cell(row, 18).value = data["qty"]
        ws.cell(row, 19).value = data["approved"]
        ws.cell(row, 20).value = data["approved"] / data["lots"] if data["lots"] else 0
        ws.cell(row, 18).number_format = '#,##0 "kg"'
        ws.cell(row, 20).number_format = "0.0%"

    status_headers = ["Status", "Receb.", "%"]
    for col, text in enumerate(status_headers, start=13):
        ws.cell(50, col).value = text
    table_style(ws, 50, 13, 53, 15, COLORS["green"])
    status_rows = ["Aprovado", "Em andamento", "Reprovado"]
    for row, label in enumerate(status_rows, start=51):
        count = summary["status"].get(label, 0)
        ws.cell(row, 13).value = label
        ws.cell(row, 14).value = count
        ws.cell(row, 15).value = count / summary["total_lots"] if summary["total_lots"] else 0
        ws.cell(row, 15).number_format = "0.0%"
    for row, color in zip(range(51, 54), [COLORS["green"], COLORS["amber"], COLORS["rose"]]):
        ws.cell(row, 13).fill = fill(color)
        ws.cell(row, 13).font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=9)

    due_headers = ["#", "Matéria", "Fornecedor", "Validade", "Dias"]
    for col, text in enumerate(due_headers, start=19):
        ws.cell(50, col).value = text
    table_style(ws, 50, 19, 60, 23, COLORS["rose"])
    today = date.today()
    for row in range(51, 61):
        idx = row - 51
        ws.cell(row, 19).value = idx + 1
        if idx < len(summary["due"]):
            item = summary["due"][idx]
            ws.cell(row, 20).value = item["material"]
            ws.cell(row, 21).value = item["supplier"]
            ws.cell(row, 22).value = item["validity"]
            ws.cell(row, 23).value = (item["validity"] - today).days
        ws.cell(row, 22).number_format = "dd/mm/yyyy"
        ws.cell(row, 23).number_format = "0"

    material_headers = ["Matéria-prima", "Receb.", "Quantidade (kg)", "Aprovados", "% Aprov.", "Fornecedores", "Rank"]
    for col, text in enumerate(material_headers, start=5):
        ws.cell(66, col).value = text
    last_material_row = 66 + len(materials)
    table_style(ws, 66, 5, max(last_material_row, 126), 11, COLORS["purple"])
    material_rows = sorted(summary["by_material"].items(), key=lambda item: item[1]["qty"], reverse=True)
    ranks = {name: idx + 1 for idx, (name, _) in enumerate(material_rows)}
    for row, (material, data) in enumerate(material_rows, start=67):
        lots = data["lots"]
        ws.cell(row, 5).value = material
        ws.cell(row, 6).value = lots
        ws.cell(row, 7).value = data["qty"]
        ws.cell(row, 8).value = data["approved"]
        ws.cell(row, 9).value = data["approved"] / lots if lots else 0
        ws.cell(row, 10).value = len(data["suppliers"])
        ws.cell(row, 11).value = ranks[material] if data["qty"] else ""
        ws.cell(row, 7).number_format = '#,##0 "kg"'
        ws.cell(row, 9).number_format = "0.0%"

    ws.conditional_formatting.add("R16:R25", DataBarRule(start_type="min", end_type="max", color=COLORS["blue"], showValue=True))
    ws.conditional_formatting.add("R33:R44", DataBarRule(start_type="min", end_type="max", color=COLORS["teal"], showValue=True))
    ws.conditional_formatting.add(f"G67:G{last_material_row}", DataBarRule(start_type="min", end_type="max", color=COLORS["purple"], showValue=True))
    ws.conditional_formatting.add(f"I67:I{last_material_row}", CellIsRule(operator="lessThan", formula=["0.8"], fill=fill(COLORS["rose2"])))
    ws.conditional_formatting.add(f"I67:I{last_material_row}", CellIsRule(operator="greaterThanOrEqual", formula=["0.95"], fill=fill(COLORS["green2"])))
    ws.conditional_formatting.add("W51:W60", CellIsRule(operator="lessThanOrEqual", formula=["30"], fill=fill(COLORS["rose2"])))
    ws.conditional_formatting.add("W51:W60", CellIsRule(operator="between", formula=["31", "90"], fill=fill(COLORS["amber2"])))

    top_chart = BarChart()
    top_chart.type = "bar"
    top_chart.title = None
    top_chart.add_data(Reference(ws, min_col=18, min_row=15, max_row=25), titles_from_data=True)
    top_chart.set_categories(Reference(ws, min_col=17, min_row=16, max_row=25))
    top_chart.height = 7.4
    top_chart.width = 18.4
    top_chart.legend = None
    top_chart.style = 10
    if top_chart.series:
        top_chart.series[0].graphicalProperties.solidFill = COLORS["blue"]
        top_chart.series[0].graphicalProperties.line.solidFill = COLORS["blue"]
    ws.add_chart(top_chart, "E14")

    month_chart = BarChart()
    month_chart.type = "col"
    month_chart.title = None
    month_chart.add_data(Reference(ws, min_col=18, min_row=32, max_row=44), titles_from_data=True)
    month_chart.set_categories(Reference(ws, min_col=16, min_row=33, max_row=44))
    month_chart.height = 7.4
    month_chart.width = 18.4
    month_chart.legend = None
    month_chart.style = 10
    if month_chart.series:
        month_chart.series[0].graphicalProperties.solidFill = COLORS["teal"]
        month_chart.series[0].graphicalProperties.line.solidFill = COLORS["teal"]
    ws.add_chart(month_chart, "E31")

    status_chart = DoughnutChart()
    status_chart.title = None
    status_chart.add_data(Reference(ws, min_col=14, min_row=50, max_row=53), titles_from_data=True)
    status_chart.set_categories(Reference(ws, min_col=13, min_row=51, max_row=53))
    status_chart.holeSize = 58
    status_chart.height = 5.8
    status_chart.width = 7.8
    status_chart.legend = None
    status_chart.style = 10
    ws.add_chart(status_chart, "F50")

    for row in ws.iter_rows(min_row=1, max_row=132, min_col=1, max_col=23):
        for cell in row:
            if cell.font == Font():
                cell.font = Font(name="Segoe UI", color=COLORS["ink"], size=9)

    for col in (16, 19):
        for row in range(16, 61):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
    for row in range(16, 127):
        for col in (7, 18, 23):
            ws.cell(row, col).alignment = Alignment(horizontal="right", vertical="center")

    ws.protection.sheet = False
    ws.protection.objects = False
    ws.protection.scenarios = False
    ws.print_area = "A1:W128"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def write_dashboard_dynamic(wb, source_sheets, records):
    for name in ("Dashboard2", "dashboard2"):
        if name in wb.sheetnames:
            wb.remove(wb[name])

    helper_last_row = HELPER_FIRST_ROW + (len(source_sheets) * ROWS_PER_MATERIAL) - 1
    supplier_names = {}
    for item in records:
        supplier = str(item["supplier"] or "").strip()
        if not supplier:
            continue
        key = normalize(supplier)
        current = supplier_names.get(key)
        if current is None or (current.isupper() and not supplier.isupper()):
            supplier_names[key] = supplier
    suppliers = sorted(supplier_names.values(), key=lambda value: normalize(value))
    supplier_last_row = max(2, 1 + len(suppliers))
    material_last_row = 66 + len(source_sheets)
    material_style_last_row = max(material_last_row, 126)

    ws = wb.create_sheet("Dashboard2", 0)
    wb.active = 0
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "E11"

    widths = {
        "A": 2.5, "B": 13, "C": 13, "D": 2.5,
        "E": 11, "F": 11, "G": 12, "H": 12, "I": 12, "J": 12, "K": 12,
        "L": 12, "M": 12, "N": 12, "O": 9, "P": 7, "Q": 17, "R": 15,
        "S": 10, "T": 10, "U": 11, "V": 10, "W": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(1, 132):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[2].height = 30

    style_area(ws, "A1:W132", COLORS["bg"])
    style_area(ws, "A1:C132", COLORS["nav"])
    style_area(ws, "D1:D132", COLORS["bg"])

    merge(ws, "B2:C2", "MATERIA PRIMA", COLORS["nav"], Font(name="Segoe UI", color="FFFFFF", bold=True, size=14), Alignment(horizontal="left", vertical="center"))
    merge(ws, "B3:C3", "Fornecedores", COLORS["nav"], Font(name="Segoe UI", color="B8C4D6", size=9), Alignment(horizontal="left", vertical="center"))
    merge(ws, "B6:C7", f'=COUNTA($E$67:$E${material_last_row})', COLORS["nav2"], Font(name="Segoe UI", color="FFFFFF", bold=True, size=22), Alignment(horizontal="center", vertical="center"), COLORS["nav2"])
    ws["B6"].number_format = "0"
    merge(ws, "B8:C8", "materias cadastradas", COLORS["nav2"], Font(name="Segoe UI", color="B8C4D6", bold=True, size=9), Alignment(horizontal="center", vertical="center"), COLORS["nav2"])
    for row, label in [(12, "Visao geral"), (14, "Fornecedores"), (16, "Mensal"), (18, "Status"), (20, "Vencimentos")]:
        merge(ws, f"B{row}:C{row}", label, COLORS["nav"], Font(name="Segoe UI", color="DDE6F3", bold=True, size=10), Alignment(horizontal="left", vertical="center", indent=1))

    merge(ws, "E2:S2", "Dashboard de Materia Prima", COLORS["bg"], Font(name="Segoe UI", color=COLORS["ink"], bold=True, size=22), Alignment(horizontal="left", vertical="center"))
    merge(ws, "E3:S3", '="Atualizado: "&TEXT(NOW(),"dd/mm/aaaa hh:mm")&"  |  Ano selecionado: "&$W$3', COLORS["bg"], Font(name="Segoe UI", color=COLORS["muted"], size=10), Alignment(horizontal="left", vertical="center"))
    merge(ws, "U2:V2", "ANO", COLORS["surface"], Font(name="Segoe UI", color=COLORS["muted"], bold=True, size=9), Alignment(horizontal="center", vertical="center"), COLORS["line"])
    merge(ws, "U3:V3", "base", COLORS["surface"], Font(name="Segoe UI", color=COLORS["muted"], size=9), Alignment(horizontal="center", vertical="center"), COLORS["line"])
    ws["W2"] = "Selecionado"
    ws["W3"] = DEFAULT_YEAR
    for cell in ("W2", "W3"):
        ws[cell].fill = fill(COLORS["surface"] if cell == "W2" else COLORS["amber2"])
        ws[cell].font = Font(name="Segoe UI", color=COLORS["ink"] if cell == "W3" else COLORS["muted"], bold=True, size=10)
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].border = border(COLORS["line"])

    card(ws, "E5:H9", "RECEBIMENTOS", f'=SUMIFS($AK$2:$AK${helper_last_row},$AH$2:$AH${helper_last_row},$W$3)', COLORS["blue"], "0")
    card(ws, "I5:L9", "QUANTIDADE TOTAL", f'=SUMIFS($AE$2:$AE${helper_last_row},$AK$2:$AK${helper_last_row},1,$AH$2:$AH${helper_last_row},$W$3)', COLORS["teal"], '#,##0 "kg"')
    card(ws, "M5:P9", "FORNECEDORES", f'=COUNTIFS($AQ$2:$AQ${supplier_last_row},">0")', COLORS["green"], "0")
    card(ws, "Q5:T9", "TAXA DE APROVACAO", f'=IFERROR(SUMIFS($AG$2:$AG${helper_last_row},$AK$2:$AK${helper_last_row},1,$AH$2:$AH${helper_last_row},$W$3)/$E$7,0)', COLORS["amber"], "0.0%")
    card(ws, "U5:W9", "MATERIAS", f'=COUNTA($E$67:$E${material_last_row})', COLORS["purple"], "0")

    panel(ws, "E12:N27", "Top 10 fornecedores por quantidade")
    panel(ws, "P12:W27", "Ranking de fornecedores")
    panel(ws, "E29:N45", "Recebimentos mensais")
    panel(ws, "P29:W45", "Detalhe mensal")
    panel(ws, "E47:K61", "Distribuicao de aprovacao")
    panel(ws, "M47:Q61", "Status de aprovacao")
    panel(ws, "S47:W61", "Proximos vencimentos")
    panel(ws, "E64:W128", "Resumo por materia-prima")

    supplier_headers = ["#", "Fornecedor", "Quantidade (kg)", "Receb.", "% Aprov.", "Materias"]
    for col, text in enumerate(supplier_headers, start=16):
        ws.cell(15, col).value = text
    table_style(ws, 15, 16, 25, 21, COLORS["blue"])
    for row in range(16, 26):
        rank = row - 15
        ws.cell(row, 16).value = rank
        ws.cell(row, 17).value = f'=IFERROR(IF(LARGE($AT$2:$AT${supplier_last_row},P{row})=0,"",INDEX($AO$2:$AO${supplier_last_row},MATCH(LARGE($AT$2:$AT${supplier_last_row},P{row}),$AT$2:$AT${supplier_last_row},0))),"")'
        ws.cell(row, 18).value = f'=IF(Q{row}="","",INDEX($AP$2:$AP${supplier_last_row},MATCH(Q{row},$AO$2:$AO${supplier_last_row},0)))'
        ws.cell(row, 19).value = f'=IF(Q{row}="","",INDEX($AQ$2:$AQ${supplier_last_row},MATCH(Q{row},$AO$2:$AO${supplier_last_row},0)))'
        ws.cell(row, 20).value = f'=IFERROR(INDEX($AR$2:$AR${supplier_last_row},MATCH(Q{row},$AO$2:$AO${supplier_last_row},0))/S{row},0)'
        ws.cell(row, 21).value = f'=IF(Q{row}="","",INDEX($AS$2:$AS${supplier_last_row},MATCH(Q{row},$AO$2:$AO${supplier_last_row},0)))'
        ws.cell(row, 18).number_format = '#,##0 "kg"'
        ws.cell(row, 20).number_format = "0.0%"

    months = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    month_headers = ["Mes", "Receb.", "Quantidade (kg)", "Aprovados", "% Aprov."]
    for col, text in enumerate(month_headers, start=16):
        ws.cell(32, col).value = text
    table_style(ws, 32, 16, 44, 20, COLORS["teal"])
    for month_no, month in enumerate(months, start=1):
        row = 32 + month_no
        ws.cell(row, 16).value = month
        ws.cell(row, 17).value = f'=COUNTIFS($AK$2:$AK${helper_last_row},1,$AH$2:$AH${helper_last_row},$W$3,$AI$2:$AI${helper_last_row},{month_no})'
        ws.cell(row, 18).value = f'=SUMIFS($AE$2:$AE${helper_last_row},$AK$2:$AK${helper_last_row},1,$AH$2:$AH${helper_last_row},$W$3,$AI$2:$AI${helper_last_row},{month_no})'
        ws.cell(row, 19).value = f'=SUMIFS($AG$2:$AG${helper_last_row},$AK$2:$AK${helper_last_row},1,$AH$2:$AH${helper_last_row},$W$3,$AI$2:$AI${helper_last_row},{month_no})'
        ws.cell(row, 20).value = f'=IFERROR(S{row}/Q{row},0)'
        ws.cell(row, 18).number_format = '#,##0 "kg"'
        ws.cell(row, 20).number_format = "0.0%"

    status_headers = ["Status", "Receb.", "%"]
    for col, text in enumerate(status_headers, start=13):
        ws.cell(50, col).value = text
    table_style(ws, 50, 13, 53, 15, COLORS["green"])
    status_rows = [
        ("Aprovado", f'=SUMIFS($AG$2:$AG${helper_last_row},$AK$2:$AK${helper_last_row},1,$AH$2:$AH${helper_last_row},$W$3)'),
        ("Em andamento", f'=COUNTIFS($AK$2:$AK${helper_last_row},1,$AH$2:$AH${helper_last_row},$W$3,$AF$2:$AF${helper_last_row},"<>",$AF$2:$AF${helper_last_row},"<>*aprovado*",$AF$2:$AF${helper_last_row},"<>*reprov*")'),
        ("Reprovado", f'=COUNTIFS($AK$2:$AK${helper_last_row},1,$AH$2:$AH${helper_last_row},$W$3,$AF$2:$AF${helper_last_row},"*reprov*")'),
    ]
    for row, (label, formula) in enumerate(status_rows, start=51):
        ws.cell(row, 13).value = label
        ws.cell(row, 14).value = formula
        ws.cell(row, 15).value = f'=IFERROR(N{row}/$E$7,0)'
        ws.cell(row, 15).number_format = "0.0%"
    for row, color in zip(range(51, 54), [COLORS["green"], COLORS["amber"], COLORS["rose"]]):
        ws.cell(row, 13).fill = fill(color)
        ws.cell(row, 13).font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=9)

    due_headers = ["#", "Materia", "Fornecedor", "Validade", "Dias"]
    for col, text in enumerate(due_headers, start=19):
        ws.cell(50, col).value = text
    table_style(ws, 50, 19, 60, 23, COLORS["rose"])
    for row in range(51, 61):
        rank = row - 50
        ws.cell(row, 19).value = rank
        ws.cell(row, 20).value = f'=IFERROR(INDEX($AA$2:$AA${helper_last_row},MATCH(SMALL($AL$2:$AL${helper_last_row},S{row}),$AL$2:$AL${helper_last_row},0)),"")'
        ws.cell(row, 21).value = f'=IFERROR(INDEX($AB$2:$AB${helper_last_row},MATCH(SMALL($AL$2:$AL${helper_last_row},S{row}),$AL$2:$AL${helper_last_row},0)),"")'
        ws.cell(row, 22).value = f'=IFERROR(INDEX($AJ$2:$AJ${helper_last_row},MATCH(SMALL($AL$2:$AL${helper_last_row},S{row}),$AL$2:$AL${helper_last_row},0)),"")'
        ws.cell(row, 23).value = f'=IF(V{row}="","",V{row}-TODAY())'
        ws.cell(row, 22).number_format = "dd/mm/yyyy"
        ws.cell(row, 23).number_format = "0"

    material_headers = ["Materia-prima", "Receb.", "Quantidade (kg)", "Aprovados", "% Aprov.", "Fornecedores", "Rank"]
    for col, text in enumerate(material_headers, start=5):
        ws.cell(66, col).value = text
    table_style(ws, 66, 5, material_style_last_row, 11, COLORS["purple"])
    for row, source in enumerate(source_sheets, start=67):
        ws.cell(row, 5).value = source["label"]
        ws.cell(row, 6).value = f'=SUMIFS($AK$2:$AK${helper_last_row},$AA$2:$AA${helper_last_row},E{row},$AH$2:$AH${helper_last_row},$W$3)'
        ws.cell(row, 7).value = f'=SUMIFS($AE$2:$AE${helper_last_row},$AK$2:$AK${helper_last_row},1,$AA$2:$AA${helper_last_row},E{row},$AH$2:$AH${helper_last_row},$W$3)'
        ws.cell(row, 8).value = f'=SUMIFS($AG$2:$AG${helper_last_row},$AK$2:$AK${helper_last_row},1,$AA$2:$AA${helper_last_row},E{row},$AH$2:$AH${helper_last_row},$W$3)'
        ws.cell(row, 9).value = f'=IFERROR(H{row}/F{row},0)'
        ws.cell(row, 10).value = f'=IF(F{row}=0,0,SUMPRODUCT(($AO$2:$AO${supplier_last_row}<>"")*(COUNTIFS($AA$2:$AA${helper_last_row},E{row},$AB$2:$AB${helper_last_row},$AO$2:$AO${supplier_last_row},$AH$2:$AH${helper_last_row},$W$3)>0)))'
        ws.cell(row, 11).value = f'=IF(G{row}>0,RANK(G{row},$G$67:$G${material_last_row},0),"")'
        ws.cell(row, 7).number_format = '#,##0 "kg"'
        ws.cell(row, 9).number_format = "0.0%"

    helper_headers = [
        "Materia",
        "Fornecedor",
        "Lote",
        "Data receb.",
        "Quantidade",
        "Status",
        "Aprovado",
        "Ano",
        "Mes",
        "Validade calc.",
        "Linha valida",
        "Vencimento ordenado",
    ]
    for col, text in enumerate(helper_headers, start=27):
        ws.cell(1, col).value = text
        ws.cell(1, col).font = Font(name="Segoe UI", bold=True)

    current_row = HELPER_FIRST_ROW
    for source in source_sheets:
        sheet_ref = quote_sheet(source["sheet"])
        h = source["header"]
        received_col = get_column_letter(h["received_date"])
        qty_col = get_column_letter(h["qty"])
        status_col = get_column_letter(h["approval"])
        validity_col = get_column_letter(h["validity"])
        supplier_col = get_column_letter(h["supplier"]) if "supplier" in h else None
        lot_col = get_column_letter(h["lot"]) if "lot" in h else None
        first_source_row = h["row"] + 1
        label = excel_text(source["label"])

        for offset in range(ROWS_PER_MATERIAL):
            source_row = first_source_row + offset
            received_ref = f"{sheet_ref}!{received_col}{source_row}"
            qty_ref = f"{sheet_ref}!{qty_col}{source_row}"
            status_ref = f"{sheet_ref}!{status_col}{source_row}"
            validity_ref = f"{sheet_ref}!{validity_col}{source_row}"
            supplier_ref = f"{sheet_ref}!{supplier_col}{source_row}" if supplier_col else '""'
            lot_ref = f"{sheet_ref}!{lot_col}{source_row}" if lot_col else '""'
            material_cell = f"AA{current_row}"
            date_cell = f"AD{current_row}"
            status_cell = f"AF{current_row}"
            year_cell = f"AH{current_row}"
            validity_cell = f"AJ{current_row}"
            valid_cell = f"AK{current_row}"

            ws.cell(current_row, 27).value = f'=IF(COUNTA({received_ref},{qty_ref},{status_ref},{validity_ref})>0,"{label}","")'
            ws.cell(current_row, 28).value = f'=IF({material_cell}="","",TRIM({supplier_ref}))'
            ws.cell(current_row, 29).value = f'=IF({material_cell}="","",{lot_ref})'
            ws.cell(current_row, 30).value = f'=IFERROR(IF(ISNUMBER({received_ref}),{received_ref},DATEVALUE({received_ref})),"")'
            ws.cell(current_row, 31).value = (
                f'=IF({material_cell}="","",IFERROR(1*{qty_ref},IFERROR(NUMBERVALUE({qty_ref},".",","),'
                f'IFERROR(NUMBERVALUE({qty_ref},",","."),0))))'
            )
            ws.cell(current_row, 32).value = f'=IF({material_cell}="","",{status_ref})'
            ws.cell(current_row, 33).value = f'=IF({material_cell}="",0,--AND(ISNUMBER(SEARCH("aprov",{status_cell})),NOT(ISNUMBER(SEARCH("reprov",{status_cell})))))'
            ws.cell(current_row, 34).value = f'=IF({date_cell}="","",YEAR({date_cell}))'
            ws.cell(current_row, 35).value = f'=IF({date_cell}="","",MONTH({date_cell}))'
            ws.cell(current_row, 36).value = (
                f'=IF({material_cell}="","",IFERROR(IF(ISNUMBER({validity_ref}),{validity_ref},'
                f'IF(ISNUMBER(SEARCH("12",{validity_ref})),EDATE({date_cell},12),'
                f'IF(ISNUMBER(SEARCH("6",{validity_ref})),EDATE({date_cell},6),'
                f'IF(ISNUMBER(SEARCH("3",{validity_ref})),EDATE({date_cell},3),DATEVALUE({validity_ref}))))),""))'
            )
            ws.cell(current_row, 37).value = f'=--({material_cell}<>"")'
            ws.cell(current_row, 38).value = f'=IF(AND({valid_cell}=1,{validity_cell}<>"",{validity_cell}>=TODAY(),{year_cell}=$W$3),{validity_cell}+ROW()/1000000,"")'
            current_row += 1

    for idx, year in enumerate(range(2022, 2031), start=2):
        ws.cell(idx, 40).value = year
    validation = DataValidation(type="list", formula1="$AN$2:$AN$10", allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(ws["W3"])

    supplier_headers_hidden = ["Fornecedor", "Quantidade", "Receb.", "Aprovados", "Materias", "Ordenacao"]
    for col, text in enumerate(supplier_headers_hidden, start=41):
        ws.cell(1, col).value = text
        ws.cell(1, col).font = Font(name="Segoe UI", bold=True)
    for row, supplier in enumerate(suppliers, start=2):
        ws.cell(row, 41).value = supplier
        ws.cell(row, 42).value = f'=SUMIFS($AE$2:$AE${helper_last_row},$AK$2:$AK${helper_last_row},1,$AB$2:$AB${helper_last_row},AO{row},$AH$2:$AH${helper_last_row},$W$3)'
        ws.cell(row, 43).value = f'=COUNTIFS($AK$2:$AK${helper_last_row},1,$AB$2:$AB${helper_last_row},AO{row},$AH$2:$AH${helper_last_row},$W$3)'
        ws.cell(row, 44).value = f'=SUMIFS($AG$2:$AG${helper_last_row},$AK$2:$AK${helper_last_row},1,$AB$2:$AB${helper_last_row},AO{row},$AH$2:$AH${helper_last_row},$W$3)'
        ws.cell(row, 45).value = f'=IF(AQ{row}=0,0,SUMPRODUCT(($E$67:$E${material_last_row}<>"")*(COUNTIFS($AA$2:$AA${helper_last_row},$E$67:$E${material_last_row},$AB$2:$AB${helper_last_row},AO{row},$AH$2:$AH${helper_last_row},$W$3)>0)))'
        ws.cell(row, 46).value = f'=IF(AP{row}>0,AP{row}+ROW()/1000000,0)'

    ws.conditional_formatting.add("R16:R25", DataBarRule(start_type="min", end_type="max", color=COLORS["blue"], showValue=True))
    ws.conditional_formatting.add("R33:R44", DataBarRule(start_type="min", end_type="max", color=COLORS["teal"], showValue=True))
    ws.conditional_formatting.add(f"G67:G{material_last_row}", DataBarRule(start_type="min", end_type="max", color=COLORS["purple"], showValue=True))
    ws.conditional_formatting.add(f"I67:I{material_last_row}", CellIsRule(operator="lessThan", formula=["0.8"], fill=fill(COLORS["rose2"])))
    ws.conditional_formatting.add(f"I67:I{material_last_row}", CellIsRule(operator="greaterThanOrEqual", formula=["0.95"], fill=fill(COLORS["green2"])))
    ws.conditional_formatting.add("W51:W60", CellIsRule(operator="lessThanOrEqual", formula=["30"], fill=fill(COLORS["rose2"])))
    ws.conditional_formatting.add("W51:W60", CellIsRule(operator="between", formula=["31", "90"], fill=fill(COLORS["amber2"])))

    top_chart = BarChart()
    top_chart.type = "bar"
    top_chart.title = None
    top_chart.add_data(Reference(ws, min_col=18, min_row=15, max_row=25), titles_from_data=True)
    top_chart.set_categories(Reference(ws, min_col=17, min_row=16, max_row=25))
    top_chart.height = 7.4
    top_chart.width = 18.4
    top_chart.legend = None
    top_chart.style = 10
    if top_chart.series:
        top_chart.series[0].graphicalProperties.solidFill = COLORS["blue"]
        top_chart.series[0].graphicalProperties.line.solidFill = COLORS["blue"]
    ws.add_chart(top_chart, "E14")

    month_chart = BarChart()
    month_chart.type = "col"
    month_chart.title = None
    month_chart.add_data(Reference(ws, min_col=18, min_row=32, max_row=44), titles_from_data=True)
    month_chart.set_categories(Reference(ws, min_col=16, min_row=33, max_row=44))
    month_chart.height = 7.4
    month_chart.width = 18.4
    month_chart.legend = None
    month_chart.style = 10
    if month_chart.series:
        month_chart.series[0].graphicalProperties.solidFill = COLORS["teal"]
        month_chart.series[0].graphicalProperties.line.solidFill = COLORS["teal"]
    ws.add_chart(month_chart, "E31")

    status_chart = DoughnutChart()
    status_chart.title = None
    status_chart.add_data(Reference(ws, min_col=14, min_row=50, max_row=53), titles_from_data=True)
    status_chart.set_categories(Reference(ws, min_col=13, min_row=51, max_row=53))
    status_chart.holeSize = 58
    status_chart.height = 5.8
    status_chart.width = 7.8
    status_chart.legend = None
    status_chart.style = 10
    ws.add_chart(status_chart, "F50")

    for col in range(24, 47):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    for row in ws.iter_rows(min_row=1, max_row=132, min_col=1, max_col=23):
        for cell in row:
            if cell.font == Font():
                cell.font = Font(name="Segoe UI", color=COLORS["ink"], size=9)

    for col in (16, 19):
        for row in range(16, 61):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
    for row in range(16, 127):
        for col in (7, 18, 23):
            ws.cell(row, col).alignment = Alignment(horizontal="right", vertical="center")

    ws.protection.sheet = False
    ws.protection.objects = False
    ws.protection.scenarios = False
    ws.print_area = "A1:W128"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def build():
    wb = openpyxl.load_workbook(INPUT, keep_vba=True)
    records, materials = collect_records(wb)
    source_sheets = find_source_sheets(wb)
    write_dashboard_dynamic(wb, source_sheets, records)
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT, len(records), len(materials)


if __name__ == "__main__":
    output, records_count, materials_count = build()
    print(f"Dashboard gerado: {output.resolve()}")
    print(f"Registros lidos: {records_count}")
    print(f"Abas de materia-prima: {materials_count}")
